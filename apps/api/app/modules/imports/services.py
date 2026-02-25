from __future__ import annotations

import csv
import hashlib
import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import psycopg
from openpyxl import load_workbook
from psycopg import sql

from app.shared.infrastructure.settings import Settings


HEADER_SANITIZE_PATTERN = re.compile(r"[^a-zA-Z0-9_]+")
FORMULA_PREFIXES = ("=", "+", "-", "@")


@dataclass(slots=True)
class ParsedSpreadsheet:
    file_format: str
    rows: list[dict[str, Any]]
    inferred_schema: list[dict[str, Any]]
    preview_rows: list[dict[str, Any]]
    row_count: int
    sheet_names: list[str]
    selected_sheet_name: str | None = None


@dataclass(slots=True)
class StoredFile:
    file_uri: str
    file_hash: str
    size_bytes: int
    absolute_path: Path


def normalize_column_name(name: str, fallback_index: int) -> str:
    normalized = unicodedata.normalize("NFKD", str(name or "")).encode("ascii", "ignore").decode("ascii")
    normalized = HEADER_SANITIZE_PATTERN.sub("_", normalized.strip().lower())
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    if not normalized:
        normalized = f"col_{fallback_index}"
    if normalized[0].isdigit():
        normalized = f"col_{normalized}"
    return normalized


def detect_file_format(filename: str) -> str:
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xlsx"):
        return "xlsx"
    raise ValueError("Only .csv and .xlsx files are supported")


def compute_sha256(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def _sanitize_formula_cell(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped.startswith(FORMULA_PREFIXES):
            if stripped[0] in {"-", "+"} and len(stripped) > 1 and stripped[1].isdigit():
                return value
            return "'" + value
    return value


def _decode_csv(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Failed to decode CSV file")


def _parse_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "1", "yes", "y", "sim"}:
            return True
        if lowered in {"false", "0", "no", "n", "nao", "não"}:
            return False
    return None


def _parse_number(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        candidate = value.strip()
        if "," in candidate and "." in candidate:
            candidate = candidate.replace(".", "").replace(",", ".")
        elif "," in candidate and "." not in candidate:
            candidate = candidate.replace(",", ".")
        try:
            return float(candidate)
        except ValueError:
            return None
    return None


def _parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return None
        try:
            return datetime.fromisoformat(candidate.replace("Z", "+00:00"))
        except ValueError:
            pass
        for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


def infer_column_type(values: list[Any]) -> str:
    non_null = [item for item in values if item not in (None, "")]
    if not non_null:
        return "string"

    bool_ok = all(_parse_bool(item) is not None for item in non_null)
    if bool_ok:
        return "bool"

    num_ok = all(_parse_number(item) is not None for item in non_null)
    if num_ok:
        return "number"

    date_hits = sum(1 for item in non_null if _parse_datetime(item) is not None)
    if date_hits / max(1, len(non_null)) >= 0.8:
        return "date"

    return "string"


def _normalize_headers(raw_headers: list[Any], max_columns: int) -> tuple[list[str], list[str]]:
    if len(raw_headers) > max_columns:
        raise ValueError(f"Spreadsheet has too many columns. Maximum allowed is {max_columns}")

    normalized_headers: list[str] = []
    display_headers: list[str] = []
    seen: dict[str, int] = {}

    for index, header in enumerate(raw_headers, start=1):
        original = str(header) if header is not None else ""
        normalized = normalize_column_name(original, index)
        if normalized in seen:
            seen[normalized] += 1
            normalized = f"{normalized}_{seen[normalized]}"
        else:
            seen[normalized] = 1
        normalized_headers.append(normalized)
        display_headers.append(original or f"Column {index}")

    return normalized_headers, display_headers


def _build_inferred_schema(
    *,
    normalized_headers: list[str],
    display_headers: list[str],
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    schema: list[dict[str, Any]] = []
    for index, (normalized, display) in enumerate(zip(normalized_headers, display_headers), start=1):
        values = [row.get(normalized) for row in rows]
        schema.append(
            {
                "index": index,
                "source_name": normalized,
                "original_name": display,
                "target_name": normalized,
                "type": infer_column_type(values),
            }
        )
    return schema


def parse_spreadsheet(
    *,
    raw: bytes,
    file_format: str,
    header_row: int,
    sheet_name: str | None,
    cell_range: str | None,
    delimiter: str | None,
    max_rows: int,
    max_columns: int,
    preview_rows: int,
) -> ParsedSpreadsheet:
    if file_format == "csv":
        csv_text = _decode_csv(raw)
        reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter or ",")
        data = list(reader)
        if not data:
            raise ValueError("CSV is empty")
        if header_row < 1 or header_row > len(data):
            raise ValueError("header_row is out of range")
        header_values = data[header_row - 1]
        normalized_headers, display_headers = _normalize_headers(header_values, max_columns=max_columns)
        content_rows = data[header_row:]
        if len(content_rows) > max_rows:
            raise ValueError(f"Spreadsheet has too many rows. Maximum allowed is {max_rows}")
        rows: list[dict[str, Any]] = []
        for raw_row in content_rows:
            row_payload: dict[str, Any] = {}
            for idx, key in enumerate(normalized_headers):
                value = raw_row[idx] if idx < len(raw_row) else None
                row_payload[key] = _sanitize_formula_cell(value)
            rows.append(row_payload)
        inferred_schema = _build_inferred_schema(
            normalized_headers=normalized_headers,
            display_headers=display_headers,
            rows=rows,
        )
        return ParsedSpreadsheet(
            file_format="csv",
            rows=rows,
            inferred_schema=inferred_schema,
            preview_rows=rows[:preview_rows],
            row_count=len(rows),
            sheet_names=["csv"],
            selected_sheet_name=None,
        )

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
    workbook_sheet_names = list(workbook.sheetnames)
    if sheet_name and sheet_name not in workbook_sheet_names:
        raise ValueError(f"Worksheet '{sheet_name}' was not found")
    selected_sheet_name = sheet_name or workbook_sheet_names[0]
    worksheet = workbook[selected_sheet_name]
    if cell_range:
        matrix = [[cell.value for cell in row] for row in worksheet[cell_range]]
    else:
        matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]

    if not matrix:
        raise ValueError("XLSX is empty")
    if header_row < 1 or header_row > len(matrix):
        raise ValueError("header_row is out of range")

    headers = matrix[header_row - 1]
    normalized_headers, display_headers = _normalize_headers(list(headers), max_columns=max_columns)
    content_rows = matrix[header_row:]
    if len(content_rows) > max_rows:
        raise ValueError(f"Spreadsheet has too many rows. Maximum allowed is {max_rows}")

    rows: list[dict[str, Any]] = []
    for row in content_rows:
        payload: dict[str, Any] = {}
        for idx, key in enumerate(normalized_headers):
            value = row[idx] if idx < len(row) else None
            payload[key] = _sanitize_formula_cell(value)
        rows.append(payload)

    inferred_schema = _build_inferred_schema(
        normalized_headers=normalized_headers,
        display_headers=display_headers,
        rows=rows,
    )
    return ParsedSpreadsheet(
        file_format="xlsx",
        rows=rows,
        inferred_schema=inferred_schema,
        preview_rows=rows[:preview_rows],
        row_count=len(rows),
        sheet_names=workbook_sheet_names,
        selected_sheet_name=selected_sheet_name,
    )


def store_uploaded_file(
    *,
    settings: Settings,
    import_id: int,
    filename: str,
    raw: bytes,
) -> StoredFile:
    size_bytes = len(raw)
    if size_bytes > settings.import_max_file_size_bytes:
        raise ValueError(f"File exceeds maximum allowed size ({settings.import_max_file_size_bytes} bytes)")

    file_hash = compute_sha256(raw)
    safe_name = re.sub(r"[^a-zA-Z0-9._-]+", "_", filename)
    storage_key = f"{import_id}/{file_hash}_{safe_name}"

    if settings.import_storage_backend == "s3":
        if not settings.import_s3_bucket:
            raise ValueError("S3 bucket is not configured")
        s3 = _create_s3_client(settings)
        object_key = f"{settings.import_s3_prefix.rstrip('/')}/{storage_key}"
        s3.put_object(Bucket=settings.import_s3_bucket, Key=object_key, Body=raw)
        return StoredFile(
            file_uri=f"s3://{settings.import_s3_bucket}/{object_key}",
            file_hash=file_hash,
            size_bytes=size_bytes,
            absolute_path=Path(),
        )

    base_path = Path(settings.import_storage_local_path)
    base_path.mkdir(parents=True, exist_ok=True)
    absolute_path = (base_path / storage_key).resolve()
    absolute_path.parent.mkdir(parents=True, exist_ok=True)
    absolute_path.write_bytes(raw)
    return StoredFile(
        file_uri=f"local://{absolute_path.as_posix()}",
        file_hash=file_hash,
        size_bytes=size_bytes,
        absolute_path=absolute_path,
    )


def load_file_from_uri(*, file_uri: str, settings: Settings) -> bytes:
    if file_uri.startswith("local://"):
        path = Path(file_uri.replace("local://", "", 1))
        return path.read_bytes()
    if file_uri.startswith("s3://"):
        bucket, key = _parse_s3_uri(file_uri)
        s3 = _create_s3_client(settings)
        response = s3.get_object(Bucket=bucket, Key=key)
        body = response.get("Body")
        if body is None:
            raise ValueError("S3 object body is empty")
        return body.read()
    raise ValueError("Unsupported file_uri protocol")


def delete_file_from_uri(*, file_uri: str, settings: Settings) -> None:
    if file_uri.startswith("local://"):
        path = Path(file_uri.replace("local://", "", 1))
        if path.exists():
            path.unlink()
        return
    if file_uri.startswith("s3://"):
        bucket, key = _parse_s3_uri(file_uri)
        s3 = _create_s3_client(settings)
        s3.delete_object(Bucket=bucket, Key=key)
        return
    raise ValueError("Unsupported file_uri protocol")


def _parse_s3_uri(file_uri: str) -> tuple[str, str]:
    without_scheme = file_uri.replace("s3://", "", 1)
    if "/" not in without_scheme:
        raise ValueError("Invalid s3 file_uri")
    bucket, key = without_scheme.split("/", 1)
    if not bucket or not key:
        raise ValueError("Invalid s3 file_uri")
    return bucket, key


def _create_s3_client(settings: Settings):
    try:
        import boto3  # type: ignore
    except ModuleNotFoundError as exc:
        raise ValueError("S3 storage backend requires boto3 dependency") from exc
    return boto3.client(
        "s3",
        region_name=settings.import_s3_region,
        endpoint_url=settings.import_s3_endpoint_url,
        aws_access_key_id=settings.import_s3_access_key_id,
        aws_secret_access_key=settings.import_s3_secret_access_key,
    )


def build_table_name(
    import_id: int,
    *,
    display_name: str | None = None,
    sheet_name: str | None = None,
    sheet_index: int | None = None,
) -> str:
    if sheet_name:
        sheet_slug = normalize_column_name(sheet_name, fallback_index=sheet_index or 1)
        raw = f"{sheet_slug}_{import_id}"
    else:
        base_slug = normalize_column_name(display_name or "import", fallback_index=import_id)
        raw = f"{base_slug}_{import_id}"

    if len(raw) <= 63:
        return raw

    if sheet_name:
        suffix = f"_{import_id}"
        max_sheet_len = max(1, 63 - len(suffix))
        return f"{sheet_slug[:max_sheet_len]}{suffix}"

    suffix = f"_{import_id}"
    max_base_len = max(1, 63 - len(suffix))
    return f"{base_slug[:max_base_len]}{suffix}"


def build_resource_id(table_name: str) -> str:
    return f"public.{table_name}"


def _to_psycopg_url(url: str) -> str:
    if url.startswith("postgresql+psycopg://"):
        return url.replace("postgresql+psycopg://", "postgresql://", 1)
    return url


def _pg_type_from_semantic(type_name: str) -> str:
    lowered = (type_name or "").lower()
    if lowered == "number":
        return "DOUBLE PRECISION"
    if lowered == "date":
        return "TIMESTAMP"
    if lowered == "bool":
        return "BOOLEAN"
    return "TEXT"


def _coerce_value(value: Any, semantic_type: str) -> Any:
    if value is None or value == "":
        return None
    lowered = (semantic_type or "").lower()
    if lowered == "number":
        return _parse_number(value)
    if lowered == "bool":
        return _parse_bool(value)
    if lowered == "date":
        return _parse_datetime(value)
    return str(value)


def create_import_table_and_load_rows(
    *,
    analytics_db_url: str,
    table_name: str,
    import_id: int,
    mapped_schema: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    error_sample_limit: int,
) -> tuple[int, list[dict[str, Any]]]:
    safe_url = _to_psycopg_url(analytics_db_url)
    error_samples: list[dict[str, Any]] = []
    inserted_rows = 0

    column_defs = [
        sql.SQL("{} {}").format(
            sql.Identifier(str(column["target_name"])),
            sql.SQL(_pg_type_from_semantic(str(column["type"]))),
        )
        for column in mapped_schema
    ]

    create_sql = sql.SQL(
        "CREATE TABLE IF NOT EXISTS {} (row_id BIGSERIAL PRIMARY KEY, import_id INTEGER NOT NULL, {})"
    ).format(sql.Identifier("public", table_name), sql.SQL(", ").join(column_defs))
    truncate_sql = sql.SQL("TRUNCATE TABLE {}").format(sql.Identifier("public", table_name))

    insert_identifiers = [sql.Identifier(str(column["target_name"])) for column in mapped_schema]
    insert_sql = sql.SQL("INSERT INTO {} (import_id, {}) VALUES ({})").format(
        sql.Identifier("public", table_name),
        sql.SQL(", ").join(insert_identifiers),
        sql.SQL(", ").join([sql.Placeholder()] * (1 + len(mapped_schema))),
    )

    batch_payload: list[tuple[Any, ...]] = []
    with psycopg.connect(safe_url) as conn:
        with conn.cursor() as cur:
            cur.execute(create_sql)
            cur.execute(truncate_sql)

            for row_index, row in enumerate(rows, start=1):
                values: list[Any] = [import_id]
                for column in mapped_schema:
                    source_name = str(column["source_name"])
                    target_name = str(column["target_name"])
                    semantic_type = str(column["type"])
                    raw_value = row.get(source_name)
                    coerced_value = _coerce_value(raw_value, semantic_type)
                    if raw_value not in (None, "") and coerced_value is None and len(error_samples) < error_sample_limit:
                        error_samples.append(
                            {
                                "row": row_index,
                                "column": target_name,
                                "value": str(raw_value),
                                "error": f"Failed to parse as {semantic_type}",
                            }
                        )
                    values.append(coerced_value)
                batch_payload.append(tuple(values))

                if len(batch_payload) >= 500:
                    cur.executemany(insert_sql, batch_payload)
                    inserted_rows += len(batch_payload)
                    batch_payload = []

            if batch_payload:
                cur.executemany(insert_sql, batch_payload)
                inserted_rows += len(batch_payload)
        conn.commit()

    return inserted_rows, error_samples
